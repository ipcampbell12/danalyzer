import os
import pandas as pd
from Helpers.columns import adjust_column_widths
from datetime import date
import calendar
import openpyxl
from openpyxl.styles import PatternFill


def delete_files_in_directory(directory_path):
   try:
     files = os.listdir(directory_path)
     for file in files:
       file_path = os.path.join(directory_path, file)
       if os.path.isfile(file_path):
         os.remove(file_path)
     print("All files deleted successfully.")
   except OSError:
     print("Error occurred while deleting files.")



def row_order_getter(grade_level):
     row_order = [
     f'All Grade {grade_level} students',
     'American Indian/Alaska Native', 'Asian', 'Black or African American',
     'Hispanic/Latino', 'Native Hawaiian/Pacific Islander','White Not Hispanic','Two or More Races', "Student of Color",'SpEd',"Section 504", 'TAG', 'EL',"Migrant", 'EconDisadvantaged','In Dual Language Program'
          ]
     
     return row_order


def cleanup_demo_sheet(ps_df, cols, y_and_n, race_cols):
    # ps_df = pd.read_excel(filename)
    all_cols = cols + y_and_n + race_cols
    remaining_cols = [col for col in all_cols if col in ps_df.columns]
    
    # Convert 'Y' and 'N' to 1 and 0 for columns in y_and_n
    for col in y_and_n + race_cols:
        if col in ps_df.columns:
            ps_df[col] = ps_df[col].map({'Y': 1, 'N': 0}).fillna(ps_df[col])
    
    # Create a copy of the DataFrame for processing
    final_df = ps_df[remaining_cols].copy()
    # print(final_df.head())
    # Calculate new columns
    final_df.loc[:, 'MoreThanOneRace'] = (final_df[race_cols].sum(axis=1) > 1).astype(int)

    final_df.loc[:, 'WhiteNotHisp'] = (
        (final_df['WhiteRaceFg'] == 1) & 
        (final_df['HispEthnicFg'].fillna(0) == 0) &
        (final_df['AmerIndianAlsknNtvRaceFg'].fillna(0) == 0) &
        (final_df['AsianRaceFg'].fillna(0) == 0) &
        (final_df['BlackRaceFg'].fillna(0) == 0) &
        (final_df['PacIslndrRaceFg'].fillna(0) == 0)
    ).astype(int)

    final_df.loc[:, 'StudentOfColorFg'] = ((final_df['WhiteNotHisp'] == 0)).astype(int)

    # print(final_df[['WhiteNotHisp', 'StudentOfColorFg']].head())
    return final_df
    # Save the updated DataFrame to Excel
    # with pd.ExcelWriter("Filtered Demographics", engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
    #     final_df.to_excel(writer, sheet_name="Filtered Demographics", index=False)


    # print("PS Demographics file filtered")
    # adjust_column_widths("Filtered Demographics")


def return_cols_from_sheet(filename):
    ps_df = pd.read_excel(filename)
    return ps_df.columns.tolist()

def return_date():
    current_month = calendar.month_name[date.today().month]
    year = date.today().year
    date_string = f"{current_month} {year}"
    return date_string

def process_date_text(rowValue):
    if pd.isna(rowValue) or str(rowValue).strip().lower() in ["", "nan"]:
        return ""
    
    # If the value is a float and ends with .0, convert to int first
    if isinstance(rowValue, float) and rowValue.is_integer():
        rowValue = str(int(rowValue))
    else:
        rowValue = str(rowValue).strip().replace('.0', '')

    if len(rowValue) == 7:
        month = "0" + rowValue[0]
        day = rowValue[1:3]
        year = rowValue[-4:]
    elif len(rowValue) == 8:
        month = rowValue[:2]
        day = rowValue[2:4]
        year = rowValue[-4:]
    else:
        print(f"Unexpected date format: {rowValue}")
        return ""
    
    return f"{month}/{day}/{year}"

def highlight_invalid_rows(excel_path, valid_record_col='Valid Record'):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Find the column index for 'Valid Record'
    header = [cell.value for cell in ws[1]]
    col_idx = header.index(valid_record_col) + 1  # openpyxl is 1-based

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        valid_record_cell = row[col_idx - 1]
        if valid_record_cell.value and str(valid_record_cell.value).strip() != "":
            for cell in row:
                cell.fill = red_fill

    wb.save(excel_path)

# After saving your DataFrame:
# discipline_df.to_excel(dest_path, index=False)
# highlight_invalid_rows(dest_path)
# adjust_column_widths(dest_path)
# print(f"Validated discipline data saved to {dest_path}")