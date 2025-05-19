from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color, Font, Border, Side

def strip_formatting(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(bold=False)
                cell.border = Border()  # No borders
    print("Formatting stripped from all sheets.")
    wb.save(path)


def adjust_column_widths(file_path, max_width=50,apply_formatting=False):
    """
    Adjust column widths in an Excel file based on content in every sheet.
    If the width exceeds max_width, text wrapping is applied.

    Args:
    - file_path (str): Path to the Excel file.
    - max_width (int): Maximum width allowed before enabling text wrapping.
    """
    wb = load_workbook(file_path)  # Load the workbook
    
    for sheet_name in wb.sheetnames:  # Loop through all sheet names
        ws = wb[sheet_name]  # Access the worksheet by name
        print(f"Adjusting column widths for sheet: {sheet_name}")
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            
            for cell in col:
                try:
                    if cell.value is not None:
                        cell_value_str = str(cell.value)  # Convert cell value to string
                        cell_length = len(cell_value_str)  # Get the length of the string
                        # print(f"Cell value: {cell.value}, Length: {cell_length}, Length Type: {type(cell_length)}")
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception as e:
                    print(f"Error in sheet '{sheet_name}', column '{column}': {e}")
            
            # Calculate adjusted width
            adjusted_width = (max_length + 2) * 1.2
            
            # Apply maximum width and enable wrapping if necessary
            # print(f"Column: {column}, Max Length: {max_length}, Adjusted Width: {adjusted_width}")
            if adjusted_width > max_width:
                ws.column_dimensions[column].width = max_width
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True)
            else:
                ws.column_dimensions[column].width = adjusted_width
        # Freeze the top row
        ws.freeze_panes = 'A2'
        if "Concerning" in sheet_name:
            ws.sheet_properties.tabColor = "FF0000"

            
    print("All column widths adjusted.")
    # strip_formatting(file_path)
    wb.save(file_path)

