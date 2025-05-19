import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from columns import adjust_column_widths
from folder_organizer import powerschool_folder, processed_yearly_numbers_folder
from pathlib import Path
import shutil

#as_of=10/01/2023
# List of schools to process
schools = ["HES", "LES", "WES", "NME", "FPMS", "VMS", "WHS", "WSHS"]

# Create a new file path for the final Excel file
final_file_path = "All_Schools_LP_Numbers.xlsx"

# Create a blank Excel writer to start with
with pd.ExcelWriter(final_file_path, engine='openpyxl') as writer:
    # Iterate through each school in the list
    for school in schools:
        # File path for each school's data
        file_path = powerschool_folder / f"{school}.xlsx"
        
        # Load data for the current school
        ps_df = pd.read_excel(file_path)
        
        # Collapse rows that are not dual language into "English Only"
        ps_df['U_StudentsUserFields.LANGUAGE_PROGRAM'] = ps_df['U_StudentsUserFields.LANGUAGE_PROGRAM'].apply(
            lambda x: "English Only" if isinstance(x, str) and "Dual" not in x else x
        )
        
        # Group by language program and calculate the total students per program
        language_group = ps_df.groupby('U_StudentsUserFields.LANGUAGE_PROGRAM')
        language_df = language_group.size().reset_index(name="Total Students").sort_values(by='Total Students', ascending=False).reset_index(drop=True)
        
        # Add a column for the school name
        language_df['School'] = school
        
        # Check if this is the first school being processed
        if writer.sheets:
            # If not the first school, append data below the existing data
            book = writer.book
            sheet = book.active
            start_row = sheet.max_row + 2  # Leave a gap between the previous school’s data
            for r_idx, row in enumerate(dataframe_to_rows(language_df, index=False, header=False), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
        else:
            # If this is the first school, write the data with headers
            language_df.to_excel(writer, index=False, sheet_name='Language Program Totals')

# Adjust column widths in the Excel file
adjust_column_widths(final_file_path)

output_file_path = Path(final_file_path)
shutil.move(output_file_path, processed_yearly_numbers_folder / output_file_path.name)

