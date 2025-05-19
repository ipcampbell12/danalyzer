import openpyxl
from openpyxl.styles import PatternFill

def highlight_large_differences(sheet, col1, col2, threshold_percentage):
    """
    Compare two columns in a single sheet and highlight cells with differences exceeding the threshold percentage.

    Parameters:
    - sheet (Worksheet): The openpyxl Worksheet object.
    - col1 (str): The letter of the first column to compare.
    - col2 (str): The letter of the second column to compare.
    - threshold_percentage (float): The percentage difference threshold for highlighting.
    """
    # Define the fill color for highlighting the differences (red)
    highlight_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Loop through the rows and compare values in the specified columns
    for row in range(2, sheet.max_row + 1):  # Starting from row 2 to skip the header
        value1 = sheet[f'{col1}{row}'].value
        value2 = sheet[f'{col2}{row}'].value

        # Ensure both values are numeric and not None
        if isinstance(value1, (int, float)) and isinstance(value2, (int, float)):
            # Avoid division by zero
            if (value1 + value2) != 0:
                # Calculate the percentage difference
                percentage_diff = abs(value1 - value2) / ((value1 + value2) / 2) * 100

                # If the percentage difference exceeds the threshold, highlight the cells
                if percentage_diff > threshold_percentage:
                    sheet[f'{col1}{row}'].fill = highlight_fill
                    sheet[f'{col2}{row}'].fill = highlight_fill

    print(f"Highlighted differences in columns {col1} and {col2} on sheet '{sheet.title}'.")

# Example usage for multiple sheets
def process_multiple_sheets(file_path, sheet_column_mapping, threshold_percentage, output_path):
    """
    Process multiple sheets in an Excel file with different column mappings for each sheet.

    Parameters:
    - file_path (str): Path to the input Excel file.
    - sheet_column_mapping (dict): A dictionary with sheet names as keys and (col1, col2) tuples as values.
    - threshold_percentage (float): The percentage difference threshold for highlighting.
    - output_path (str): Path to save the modified Excel file.
    """
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)

    # Loop through the sheet-column mappings
    for sheet_name, columns in sheet_column_mapping.items():
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            col1, col2 = columns
            highlight_large_differences(sheet, col1, col2, threshold_percentage)
        else:
            print(f"Sheet '{sheet_name}' not found in workbook.")

    # Save the modified Excel file
    wb.save(output_path)
    print(f"File saved to {output_path}")


