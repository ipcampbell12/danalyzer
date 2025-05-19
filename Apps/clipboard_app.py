import pyperclip
from openpyxl import load_workbook
from datetime import datetime

def process_clipboard_to_excel_and_text(input_path,output_folder):
    # Get the clipboard content
    clipboard_content = pyperclip.paste()

    # Determine the number of columns in the clipboard content
    rows = clipboard_content.split('\n')
    columns = rows[0].split('\t')

    # Load the existing Excel workbook
    # file_path = r"C:\Users\inpcampbell\Desktop\Email Imports.xlsx"
    wb = load_workbook(input_path)

    # Select the appropriate worksheet based on the number of columns
    if len(columns) == 7:
        ws = wb['Full Import']
        text_file_name = f"{datetime.now().strftime('%m-%d-%y')} Email Imports A.txt"
    elif len(columns) == 2:
        ws = wb['Email Import']
        text_file_name = f"{datetime.now().strftime('%m-%d-%y')} Email Imports B.txt"
    else:
        raise ValueError("Clipboard content does not match expected column counts (2 or 7)")

    # Paste the clipboard content into the selected worksheet starting from cell A2
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, cell_value in enumerate(row.split('\t'), start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # Save the workbook
    wb.save(input_path)

    print(f"Clipboard content has been pasted into the appropriate worksheet in {input_path}")

    # Create a tab-delimited text file for the selected worksheet
    # output_folder = r"C:\Users\inpcampbell\Desktop\Processing Output\Email Imports"  # Replace with your desired output folder
    text_file_path = f"{output_folder}\\{text_file_name}"

    with open(text_file_path, 'w', encoding='utf-8') as f:
        for row in ws.iter_rows(values_only=True):
            row_data = '\t'.join([str(cell) if cell is not None else '' for cell in row])
            f.write(row_data + '\n')

    print(f"{text_file_name} with {columns} columns has been created in {output_folder}")

    # Clear the worksheet except for the headers
    clear_worksheet_except_headers(ws)

    # Save the workbook again after clearing the data
    wb.save(input_path)

def clear_worksheet_except_headers(ws):
    """Clear all data in the worksheet except for the headers."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

# Call the function to execute the process
# process_clipboard_to_excel_and_text()