import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font,Alignment
import os
from Apps.Map_Analyzer.create_pivot_sheets import create_pivot_sheets
from Helpers.map_config import rows_to_highlight,other_cols, race_cols, y_and_n_cols
from Helpers.columns import adjust_column_widths
from Helpers.other_funcs import cleanup_demo_sheet, return_date
from Helpers.map_config import columns_to_keep
import subprocess
from Apps.Map_Analyzer.merge_map_and_ps import combine_excel_workbooks


def filter_out_columns(df, columns_to_keep):
    return df[columns_to_keep]


def process_map_growth_data(fall_map_input,winter_map_input,spring_map_input,ps_file,lang_file,dest_folder):
    output_merged_file = dest_folder / f'24-25 Combined MAP Growth Data - {return_date()}.xlsx'
    output_pivot_file = dest_folder / f'Disaggregated Growth Pivots - {return_date()}.xlsx'
    # COMMENT OUT THIS CODE IF YOU ALREADY MADE THE MERGED SHEET
 
    combine_excel_workbooks([fall_map_input,winter_map_input,spring_map_input], output_merged_file, columns_to_keep, ps_file,lang_file)

    map_df = pd.read_excel(output_merged_file, index_col=None)
    print(map_df.head())
    print(map_df.shape)
    create_pivot_sheets(map_df, output_pivot_file)

    wb = load_workbook(output_pivot_file)
    sheets = wb.sheetnames

    # Iterate over each sheet
    for sheet in sheets:
        print(f"The sheet name is {sheet}")

        ws = wb[sheet]

        # 1. Merge the first row across 5 columns and add the sheet title in bold and centered
        sheet_title = sheet  # Use the sheet name as the title
        ws["A1"] = sheet_title

        # Set the font to bold
        ws["A1"].font = Font(bold=True)

        # Center the text horizontally and vertically
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # 2. Highlight rows based on 'rows_to_highlight'
        for row in ws.iter_rows(min_row=2):  # Start from row 2 to avoid overwriting the title
            for cell in row[:5]:  # Limit to the first 5 columns of the row
                if cell.value in rows_to_highlight:
                    # Apply the color fill to the first five columns
                    for highlight_cell in row[:5]:  # Loop over the first 5 columns
                        highlight_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Light yellow

    # Save the changes to the workbook
    wb.save(output_pivot_file)

    # Adjust column widths (assuming adjust_column_widths function exists)
    adjust_column_widths(output_pivot_file)

    wb.close()

    # Open the final spreadsheet in Excel
    subprocess.run(["start", "excel", output_pivot_file], shell=True)



