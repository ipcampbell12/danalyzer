import pandas as pd
from Helpers.map_config import column_order, ethnicity_columns, other_columns, fluency_level_cols, race_cols
from Helpers.other_funcs import row_order_getter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


def create_ordered_pivot_table(df, term, language, grade_level="KG", row_order=None, column_order=None, ethnicity_cols=None, other_cols=None, race_cols=None, fluency_level_col=None):
    # Filter the dataframe based on grade_level, term, and language
    filtered_df = df[(df["EnrlGrdCd"] == grade_level) & 
                     (df["TermName"].str.contains(term, na=False)) & 
                     (df["TestLanguage"] == language)].copy()
    # print(filtered_df.head(15))
    # Check if filtered_df is empty
    if filtered_df.empty:
        print(f"No data found for grade_level={grade_level}, term={term}, language={language}")
        return pd.DataFrame()  # Return an empty DataFrame if no data matches the criteria

    # Map AchievementQuintile to risk levels
    risk_mapping = {
        'Exceeds Expectation': 'Low Risk',
        'Meets Expectation': 'Low Risk',
        'Approaching Expectation': 'Some Risk',
        'Below Expectation': 'High Risk'
    }

    # Apply the mapping to the fluency level column
    if fluency_level_col in filtered_df.columns:
        filtered_df['RiskLevel'] = filtered_df[fluency_level_col].map(risk_mapping)
    else:
        print(f"Column {fluency_level_col} not found in dataframe.")
        return pd.DataFrame()

    # Create MoreThanOneRace column
    if all(col in filtered_df.columns for col in race_cols):
        filtered_df['MoreThanOneRace'] = filtered_df[race_cols].sum(axis=1) >= 2
        filtered_df['MoreThanOneRace'] = filtered_df['MoreThanOneRace'].astype(int)
    else:
        print("Error: Required columns for 'MoreThanOneRace' calculation are missing.")

    # Ensure 'WhiteNotHisp' column is created
    if 'WhiteRaceFg' in filtered_df.columns and 'HispEthnicFg' in filtered_df.columns:
        filtered_df['WhiteNotHisp'] = (filtered_df['WhiteRaceFg'] == 1) & (filtered_df['HispEthnicFg'] == 0)
        filtered_df['WhiteNotHisp'] = filtered_df['WhiteNotHisp'].astype(int)
    else:
        print("Error: Required columns for 'WhiteNotHisp' calculation are missing.")

    # Define melt columns, including 'WhiteNotHisp' and 'MoreThanOneRace'
    melt_columns = ethnicity_cols + other_cols + ['WhiteNotHisp', 'MoreThanOneRace']

    # Melt the filtered dataframe for categorical columns including 'WhiteNotHisp' and 'MoreThanOneRace'
    melted_categorical = filtered_df.melt(
        id_vars=["RiskLevel"],
        value_vars=melt_columns,
        var_name='Category',
        value_name='Value'
    )

    # Replace long category names with aliases
    melted_categorical['Category'] = melted_categorical['Category'].replace({
        'AmerIndianAlsknNtvRaceFg': 'American Indian/Alaska Native',
        'AsianRaceFg': 'Asian',
        'BlackRaceFg': 'Black or African American',
        'HispEthnicFg': 'Hispanic/Latino',
        'PacIslndrRaceFg': 'Native Hawaiian/Pacific Islander',
        'WhiteRaceFg': 'White',
        'StudentOfColorFg':"Student of Color",
        'SpEdFg': 'SpEd',
        'TAGFg': 'TAG',
        'ELFg': 'EL',
        "Sect504Fg":"Section 504",
        "MigrntEdFg": "Migrant",
        'EconDsvntgFg': 'EconDisadvantaged',
        'MoreThanOneRace': 'Two or More Races',
        'WhiteNotHisp': 'White Not Hispanic',
        "InDualFg":"In Dual Language Program"
    })

    # Create a pivot table for categorical columns with sum aggregation
    pivot_categorical = melted_categorical.pivot_table(
        index='Category',
        columns='RiskLevel',
        values='Value',
        aggfunc='sum',
        fill_value=0
    )

    # Convert pivot table values to integers
    pivot_categorical = pivot_categorical.astype(int)

    # Calculate total students for the entered grade level by risk level
    grade_level_risk = filtered_df.pivot_table(
        index=["Grade"],
        columns='RiskLevel',
        aggfunc='size',
        fill_value=0
    )

    # Rename the index from grade_level to 'Total Students'
    grade_level_risk.index = [f'All Grade {grade_level} students']
    
    # Ensure all risk levels are present
    for col in ["Low Risk", "Some Risk", "High Risk"]:
        if col not in grade_level_risk.columns:
            grade_level_risk[col] = 0

    # Concatenate 'Total Students' row with pivot_categorical
    pivot_combined = pd.concat([grade_level_risk, pivot_categorical])

    # Ensure all categories in row_order are present in pivot_combined index
    if row_order:
        row_order_filtered = [cat for cat in row_order if cat in pivot_combined.index]
        pivot_combined = pivot_combined.loc[row_order_filtered]

    # Ensure all columns in column_order are present in pivot_combined columns
    if column_order:
        column_order_filtered = [col for col in column_order if col in pivot_combined.columns]
        pivot_combined = pivot_combined[column_order_filtered]

    # Add a total column
    pivot_combined["Total"] = pivot_combined.sum(axis=1)

    return pivot_combined

def create_pivot_sheets(map_df,path_to_save):
    print("Creating pivot tables")
    grades = ["KG", 1]
    terms = ["Fall", "Winter","Spring"]
    languages = ["English", "Spanish"]

    wb = Workbook()

    for language in languages:
        for term in terms:
            for grade in grades:
                sheet_name = f"{language}_{term}_{grade}"
                sheet = wb.create_sheet(title=sheet_name)
                current_row = 1  # Start from the first row

                for fluency_level_col in fluency_level_cols:
                    row_order_list = row_order_getter(grade)
                    print(f"grade {grade} and term {term} and {language} are being passed to the pivot table function ")
                    # Create the pivot table
                    result_pivot_df = create_ordered_pivot_table(
                        map_df, term, language, grade, row_order=row_order_list, column_order=column_order,
                        ethnicity_cols=ethnicity_columns, other_cols=other_columns, race_cols=race_cols,
                        fluency_level_col=fluency_level_col
                    )

                    if result_pivot_df.empty:
                        continue  # Skip if the pivot table is empty

                    # Add a title for each pivot table
                    title = f"Term: {term}, Grade {grade}, {language}, {fluency_level_col}"
                    sheet.cell(row=current_row, column=1, value=title).font = Font(bold=True)
                    current_row += 1

                    # Write the pivot table to the sheet
                    for r_idx, row in enumerate(dataframe_to_rows(result_pivot_df, index=True, header=True), start=current_row):
                        for c_idx, value in enumerate(row, start=1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)

                    current_row += len(result_pivot_df) + 3  # Move to the next row, add space between tables

                print(f"Pivot tables written to {sheet_name} successfully.")

    # Remove the default sheet created by openpyxl
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    # Save the workbook
    wb.save(path_to_save)
    print("Excel file with pivot tables created successfully.")
