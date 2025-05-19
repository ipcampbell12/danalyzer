from Helpers.config import ps_template_cols_array
import pandas as pd
from Helpers.columns import adjust_column_widths
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import subprocess
import os
from datetime import date
import calendar

cols_to_kill = [
    'First Name', 'DateStored', 'Middle Name', 'Last Name',
    'Course Number', 'Credit Type', 'Teacher Name', 'Enroll Status', 'Gender',
    'Schoolname', 'Schoolid', 'Hist Grade Level', 'Lepfg',
    'Migrntedfg', 'Sect504fg', 'Spedfg', 'Current Schoolid', 'Termid', 'Datestored'
]

# math_path = "8th Grade Algebra Grades.xlsx"
# final_path = "Math Upload.xlsx"
def determine_half(StoreCode):
        return "S1" if StoreCode in ["Q1", "Q2"] else "S2"

def determine_gpa_points(grade):
    if grade == "A":
        return 4.0
    elif grade == "B":
        return 3.0
    elif grade == "C":
        return 2.0
    elif grade == "D":
        return 1.0
    elif grade == "F":
        return 0.0
    else:
        return None

def analyze_math_grades(credits_df):
    # Ensure we're working with a copy
    df = credits_df.copy()

    # Only keep relevant StoreCodes (Q1-Q4)
    df = df[df["StoreCode"].isin(["Q1", "Q2", "Q3", "Q4"])]

    df["Half"] = df["StoreCode"].apply(determine_half)

    results = []

    # Group by student_number
    for student_number, group in df.groupby("student_number"):
        for half in ["S1", "S2"]:
            half_group = group[group["Half"] == half]

            if set(half_group["StoreCode"]) >= {"Q1", "Q2"} and half == "S1":
                row = half_group.iloc[0].copy()
                row["StoreCode"] = "S1"
                row["EarnedCrHrs"] = 0.5 if not any(half_group["Grade"] == "F") else 0.0
                # Retain the existing Grade value
                row["Grade"] = row["Grade"]
                results.append(row)

            elif set(half_group["StoreCode"]) >= {"Q3", "Q4"} and half == "S2":
                row = half_group.iloc[0].copy()
                row["StoreCode"] = "S2"
                row["EarnedCrHrs"] = 0.5 if not any(half_group["Grade"] == "F") else 0.0
                # Retain the existing Grade value
                row["Grade"] = row["Grade"]
                results.append(row)

    if results:
        final_df = pd.DataFrame(results)
        # Make sure EarnedCrHrs is exactly 0.0 or 0.5
        return final_df[credits_df.columns]  # Return in original column order
    else:
        return pd.DataFrame(columns=credits_df.columns)


def highlight_rows_by_student(wb, sheet_name):
    ws = wb[sheet_name]

    # Define fill colors
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Find the column indices for "student_number", "[students]lastFirst", and "Grade"
    student_number_col_idx = None
    last_first_col_idx = None
    grade_col_idx = None

    for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        if col[0].value == "student_number":
            student_number_col_idx = col_idx
        elif col[0].value == "[students]lastFirst":
            last_first_col_idx = col_idx
        elif col[0].value == "Grade":
            grade_col_idx = col_idx

    if not all([student_number_col_idx, last_first_col_idx, grade_col_idx]):
        print("Error: Missing required columns for highlighting.")
        return

    # Group rows by "student_number" and "[students]lastFirst"
    student_groups = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        student_key = (row[student_number_col_idx - 1].value, row[last_first_col_idx - 1].value)
        if student_key not in student_groups:
            student_groups[student_key] = []
        student_groups[student_key].append(row)

    # Apply highlighting
    for group_rows in student_groups.values():
        # Check if any row in the group has an "F" in the Grade column
        has_f_grade = any(row[grade_col_idx - 1].value == "F" for row in group_rows)
        if has_f_grade:
            for row in group_rows:
                for cell in row:
                    # Highlight rows with "F" in red, others in yellow
                    if row[grade_col_idx - 1].value == "F":
                        cell.fill = red_fill
                    else:
                        cell.fill = yellow_fill


def process_math_credits(math_path, dest_folder, ps_template_cols_array=ps_template_cols_array, cols_to_kill=cols_to_kill):


    # Read in the data
    math_df = pd.read_excel(math_path)
    

    # Drop unnecessary columns if they exist
    math_df.drop(columns=[col for col in cols_to_kill if col in math_df.columns], inplace=True, errors='ignore')
  

    # Rename fields if needed
    math_df.rename(columns={
        "[students]student_number": "student_number",
        "[students]grade_level": "Grade_Level"
    }, inplace=True)

    # Create the output dataframe with the right columns
    output_df = pd.DataFrame()

    for col_name, default_value in ps_template_cols_array:
        if col_name in math_df.columns:
            output_df[col_name] = math_df[col_name]
        else:
            output_df[col_name] = [default_value] * len(math_df)

    # Set hardcoded values
    output_df["SchoolName"] = "Woodburn High School"
    output_df["Schoolid"] = 800
    output_df["Course_Name"] = "Geometry and Data Reasoning"
    output_df["Course_Number"] = "MA101"
    output_df["StoreCode"] = math_df["StoreCode"].apply(lambda x: "S1" if x in ["Q1", "Q2"] else "S2")
    output_df["Percent"] = math_df["Percent"].astype(int)
    output_df["GPA Points"] = math_df["Grade"].apply(determine_gpa_points)  
    # import_df = analyze_math_grades(output_df)
    # import_df = import_df[import_df['EarnedCrHrs'] != 0.0]
   
    # Remove duplicates from math_df for the merge
    math_df_unique = math_df[["student_number", "[students]lastFirst",'StoreCode']].drop_duplicates(subset="student_number")

    # Merge only the unique rows back into output_df
    output_df = output_df.merge(
        math_df_unique[["student_number", "[students]lastFirst"]],
        on="student_number",
        how="left"
    )

    # Merge only the unique rows back into import_df
    # import_df = import_df.merge(
    #     math_df_unique[["student_number", "[students]lastFirst"]],
    #     on="student_number",
    #     how="left"
    # )
  
    last_first_col = output_df.pop("[students]lastFirst")
    output_df.insert(0, "[students]lastFirst", last_first_col)
    output_df = output_df.sort_values(by=["[students]lastFirst","StoreCode"])

    # last_first_col = import_df.pop("[students]lastFirst")
    # import_df.insert(0, "[students]lastFirst", last_first_col)
    # import_df = import_df.sort_values(by="[students]lastFirst")
    current_month = calendar.month_name[date.today().month]
    year = date.today().year
    day = date.today().day
   
    final_path = os.path.join(dest_folder, f"{current_month}-{day}-{year} Math Upload.xlsx")
    # Save to Excel with multiple sheets
    with pd.ExcelWriter(final_path, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='Pre Processed Credits', index=False)
        # import_df.to_excel(writer, sheet_name='Final Processed Credits', index=False)

    # Adjust column widths for both sheets
    adjust_column_widths(final_path)
 

    # Highlight rows where Grade is 'F'
    wb = load_workbook(final_path)
    highlight_rows_by_student(wb, 'Pre Processed Credits')  # Apply highlighting using the new function

    wb.save(final_path)

    # Launch Excel file
    subprocess.run(["start", "excel", final_path], shell=True)


# process_math_credits(math_path, final_path, ps_template_cols_array, cols_to_kill)

